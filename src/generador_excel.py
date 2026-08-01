import logging
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from src.config import (
    HOJA_CONTROL,
    HOJA_DATOS_SALIDA,
    RUTA_INFORME_EXCEL,
    SALIDA_DIR,
)


logger = logging.getLogger(__name__)


COLOR_VERDE_OSCURO = "1E8449"
COLOR_VERDE_CLARO = "E9F7EF"
COLOR_BLANCO = "FFFFFF"
COLOR_BORDE = "AAB7B8"

# ==========================================================
# COLORES DE ESTADOS ANS
# ==========================================================

COLOR_VENCIDO = "FF1F1F"       # Rojo
COLOR_ALERTA = "FFD426"        # Amarillo
COLOR_A_TIEMPO = "8BCF4A"      # Verde
COLOR_INMEDIATO = "F39C12"     # Naranja
COLOR_HV = "3498DB"            # Azul
COLOR_FACTIBILIDAD = "17A589"  # Azul verdoso
COLOR_PENDIENTE = "647687"     # Gris azulado

COLOR_TEXTO_ESTADO = "000000"
COLOR_TEXTO_PENDIENTE = "FFFFFF"


ANCHOS_COLUMNAS = {
    "A": 16,   # ID_ORDEN
    "B": 15,   # FECHA_ORDEN
    "C": 45,   # DIRECCION
    "D": 33,   # PROPIETARIO
    "E": 10,   # ZONA
    "F": 15,   # MUNICIPIO
    "G": 24,   # DESC_MUNICIPIO
    "H": 17,   # REGION_ORIGEN
    "I": 18,   # TIPO
    "J": 16,   # DIAS_PACTADOS
    "K": 20,   # FECHA_LIMITE_ANS
    "L": 22,   # DIAS_TRANSCURRIDOS
    "M": 18,   # DIAS_RESTANTES
    "N": 25,   # ESTADO
    "O": 110,  # OBSERVACION
}


class ErrorGeneracionExcel(Exception):
    """Error controlado al generar el informe Excel."""


def crear_dataframe_control(
    controles_archivos: list[dict],
    control_transformacion: dict,
) -> pd.DataFrame:
    """
    Construye una hoja de trazabilidad del proceso.
    """

    registros: list[dict] = []

    for control in controles_archivos:

        registros.append(
            {
                "TIPO": "ARCHIVO",
                "ELEMENTO": control["REGION"],
                "DETALLE": (
                    f"{control['REGISTROS_VALIDOS']} "
                    f"registros procesados correctamente"
                ),
            }
        )

    registros.append(
        {
            "TIPO": "RESUMEN",
            "ELEMENTO": "Total consolidado",
            "DETALLE": (
                f"{control_transformacion['REGISTROS_CONSOLIDADOS']} registros"
            ),
        }
    )

    return pd.DataFrame(
        registros
    )


def aplicar_diseno_hoja_datos(
    ruta_archivo: Path,
) -> None:
    """
    Convierte DATOS_ANS en una tabla profesional.
    """

    libro = load_workbook(
        ruta_archivo
    )

    hoja = libro[
        HOJA_DATOS_SALIDA
    ]

    hoja.freeze_panes = "A2"
    hoja.sheet_view.showGridLines = False

    borde_fino = Side(
        style="thin",
        color=COLOR_BORDE,
    )

    # ======================================================
    # ENCABEZADOS
    # ======================================================

    for celda in hoja[1]:

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor=COLOR_VERDE_OSCURO,
        )

        celda.font = Font(
            color=COLOR_BLANCO,
            bold=True,
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        celda.border = Border(
            left=borde_fino,
            right=borde_fino,
            top=borde_fino,
            bottom=borde_fino,
        )

    hoja.row_dimensions[1].height = 30

    ultima_fila = hoja.max_row
    ultima_columna = hoja.max_column

    # ======================================================
    # TABLA ESTRUCTURADA
    # ======================================================

    if ultima_fila >= 2:

        if "TablaDatosANS" in hoja.tables:
            del hoja.tables["TablaDatosANS"]

        referencia = (
            f"A1:"
            f"{hoja.cell(1, ultima_columna).column_letter}"
            f"{ultima_fila}"
        )

        tabla = Table(
            displayName="TablaDatosANS",
            ref=referencia,
        )

        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        hoja.add_table(
            tabla
        )

    # ======================================================
    # ANCHOS DE COLUMNAS
    # ======================================================

    for columna, ancho in ANCHOS_COLUMNAS.items():
        hoja.column_dimensions[columna].width = ancho

    # ======================================================
    # FORMATO DE FILAS
    # ======================================================

    for fila in range(
        2,
        ultima_fila + 1,
    ):

        # ID_ORDEN como texto.
        hoja.cell(
            row=fila,
            column=1,
        ).number_format = "@"

        # FECHA_ORDEN.
        hoja.cell(
            row=fila,
            column=2,
        ).number_format = "dd/mm/yyyy"

        # MUNICIPIO como texto.
        hoja.cell(
            row=fila,
            column=6,
        ).number_format = "@"

        # FECHA_LIMITE_ANS ahora está en K.
        hoja.cell(
            row=fila,
            column=11,
        ).number_format = "dd/mm/yyyy"

        for columna in range(
            1,
            ultima_columna + 1,
        ):

            celda = hoja.cell(
                row=fila,
                column=columna,
            )

            # Dirección y propietario.
            if columna in {
                    3,
                    4,
                }:
                    celda.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=False,
                    )

            # Observación.
            elif columna == 15:
                celda.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
            

            # Columnas centradas.
            elif columna in {
                2,   # FECHA_ORDEN
                5,   # ZONA
                6,   # MUNICIPIO
                8,   # REGION_ORIGEN
                9,   # TIPO
                10,  # DIAS_PACTADOS
                11,  # FECHA_LIMITE_ANS
                12,  # DIAS_TRANSCURRIDOS
                13,  # DIAS_RESTANTES
                14,  # ESTADO
            }:
                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                )

            else:
                celda.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False,
                )

        hoja.row_dimensions[fila].height = 22

    # ======================================================
    # COLORES DIRECTOS PARA ESTADO
    # ======================================================

    for fila in range(
        2,
        ultima_fila + 1,
    ):

        # ESTADO ahora está en la columna N = 14.
        celda_estado = hoja.cell(
            row=fila,
            column=14,
        )

        estado = str(
            celda_estado.value or ""
        ).strip().upper()

        if estado == "VENCIDO":

            color_fondo = COLOR_VENCIDO
            color_texto = COLOR_TEXTO_ESTADO

        elif estado == "ALERTA":

            color_fondo = COLOR_ALERTA
            color_texto = COLOR_TEXTO_ESTADO

        elif estado == "A TIEMPO":

            color_fondo = COLOR_A_TIEMPO
            color_texto = COLOR_TEXTO_ESTADO

        elif estado == "INMEDIATO":

            color_fondo = COLOR_INMEDIATO
            color_texto = COLOR_TEXTO_ESTADO

        elif estado == "HV":

            color_fondo = COLOR_HV
            color_texto = COLOR_BLANCO

        elif estado == "FACTIBILIDAD":

            color_fondo = COLOR_FACTIBILIDAD
            color_texto = COLOR_BLANCO

        elif estado in {
            "SIN FECHA",
            "PENDIENTE CONFIGURACIÓN",
        }:

            color_fondo = COLOR_PENDIENTE
            color_texto = COLOR_TEXTO_PENDIENTE

        else:

            color_fondo = None
            color_texto = COLOR_TEXTO_ESTADO

        if color_fondo:

            celda_estado.fill = PatternFill(
                fill_type="solid",
                fgColor=color_fondo,
            )

        celda_estado.font = Font(
            color=color_texto,
            bold=True,
        )

        celda_estado.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False,
        )

    libro.save(
        ruta_archivo
    )
def aplicar_diseno_hoja_control(
    ruta_archivo: Path,
) -> None:
    """
    Aplica formato a la hoja de control.
    """

    libro = load_workbook(
        ruta_archivo
    )

    hoja = libro[
        HOJA_CONTROL
    ]

    hoja.freeze_panes = "A2"
    hoja.sheet_view.showGridLines = False

    borde = Side(
        style="thin",
        color=COLOR_BORDE,
    )

    for celda in hoja[1]:

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor=COLOR_VERDE_OSCURO,
        )

        celda.font = Font(
            color=COLOR_BLANCO,
            bold=True,
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        celda.border = Border(
            left=borde,
            right=borde,
            top=borde,
            bottom=borde,
        )

    hoja.column_dimensions["A"].width = 20
    hoja.column_dimensions["B"].width = 35
    hoja.column_dimensions["C"].width = 80

    libro.save(
        ruta_archivo
    )

def generar_informe_excel(
    dataframe: pd.DataFrame,
    controles_archivos: list[dict],
    control_transformacion: dict,
) -> Path:
    """
    Genera el único informe unificado.
    """

    try:
        SALIDA_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        control_dataframe = crear_dataframe_control(
            controles_archivos,
            control_transformacion,
        )

        with pd.ExcelWriter(
            RUTA_INFORME_EXCEL,
            engine="openpyxl",
            mode="w",
        ) as escritor:

            dataframe.to_excel(
                escritor,
                sheet_name=HOJA_DATOS_SALIDA,
                index=False,
            )

            control_dataframe.to_excel(
                escritor,
                sheet_name=HOJA_CONTROL,
                index=False,
            )

        aplicar_diseno_hoja_datos(
            RUTA_INFORME_EXCEL
        )

        aplicar_diseno_hoja_control(
            RUTA_INFORME_EXCEL
        )

        logger.info(
            "Informe Excel generado: %s",
            RUTA_INFORME_EXCEL,
        )

        return RUTA_INFORME_EXCEL

    except Exception as error:
        logger.exception(
            "No fue posible generar el informe Excel."
        )

        raise ErrorGeneracionExcel(
            f"No fue posible generar el informe: {error}"
        ) from error