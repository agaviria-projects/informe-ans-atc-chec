from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.config import BASE_DIR
from src.procesador_redes import procesar_exporte_redes


# ============================================================
# RUTAS
# ============================================================

CARPETA_SALIDA_REDES = BASE_DIR / "salida_redes"

NOMBRE_ARCHIVO_SALIDA = "INFORME_ANS_REDES.xlsx"
NOMBRE_HOJA = "DATOS_ANS_REDES"


# ============================================================
# COLUMNAS DEL INFORME
# ============================================================

COLUMNAS_SALIDA = [
    "NUMERO_PROCESO",
    "PROCESO",
    "D_PROCESO",
    "PRO_CLASIFICACION",
    "PRO_D_CLASIFICACION",
    "PRO_FECHA_SISTEMA_CREACION",
    "PRO_FECHA_VENCIMIENTO",
    "CLIENTE_ID",
    "CLI_NOMBRE",
    "CLI_DIRECCION",
    "CLI_D_CODIGO_AREA",
    "CLI_D_MUNICIPIO",
    "CLI_D_BARRIO",
    "REV_D_TIPO",
    "REV_ESTADO",
    "REV_RESPONSABLE",
    "REV_COMENTARIO",
    "CODIGO_UBIC_TRANSFORMADOR",
    "DIAS_CONTRACTUALES",
    "FECHA_LIMITE_ANS",
    "DIAS_TRANSCURRIDOS",
    "DIAS_PARA_INICIAR_ALERTA",
    "DIAS_RESTANTES",
    "ESTADO",
]


# ============================================================
# ANCHOS
# ============================================================

ANCHOS_COLUMNAS = {
    "A": 18, "B": 12, "C": 32, "D": 20, "E": 22, "F": 26,
    "G": 22, "H": 16, "I": 28, "J": 35, "K": 24, "L": 24,
    "M": 24, "N": 22, "O": 14, "P": 24, "Q": 40, "R": 30,
    "S": 22, "T": 26, "U": 22, "V": 26, "W": 18, "X": 18,
}


# ============================================================
# ESTILOS
# ============================================================

RELLENO_ENCABEZADO = PatternFill(
    fill_type="solid",
    fgColor="008D46",
)

FUENTE_ENCABEZADO = Font(
    color="FFFFFF",
    bold=True,
)

BORDE_FINO = Side(
    style="thin",
    color="D9E0E5",
)

BORDE_CELDA = Border(
    left=BORDE_FINO,
    right=BORDE_FINO,
    top=BORDE_FINO,
    bottom=BORDE_FINO,
)

# Fondos de ESTADO
RELLENO_ALERTA = PatternFill(
    fill_type="solid",
    fgColor="FFD426",  # amarillo
)

RELLENO_VENCIDO = PatternFill(
    fill_type="solid",
    fgColor="FF1F1F",  # rojo
)

RELLENO_ALERTA_CERO = PatternFill(
    fill_type="solid",
    fgColor="F39C12",  # naranja
)

RELLENO_A_TIEMPO = PatternFill(
    fill_type="solid",
    fgColor="8BCF4A",  # verde
)

FUENTE_ESTADO = Font(
    color="000000",
    bold=True,
)


# ============================================================
# PREPARACIÓN DE FECHAS
# ============================================================

def _preparar_fechas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["PRO_FECHA_SISTEMA_CREACION"] = pd.to_datetime(
        df["PRO_FECHA_SISTEMA_CREACION"],
        errors="coerce",
        dayfirst=True,
    )

    df["PRO_FECHA_VENCIMIENTO"] = pd.to_datetime(
        df["PRO_FECHA_VENCIMIENTO"],
        errors="coerce",
        dayfirst=True,
    ).dt.normalize()

    df["FECHA_LIMITE_ANS"] = pd.to_datetime(
        df["FECHA_LIMITE_ANS"],
        errors="coerce",
        dayfirst=True,
    )

    return df


# ============================================================
# FORMATO EXCEL
# ============================================================

def _aplicar_formato(ruta_archivo: Path) -> None:
    wb = load_workbook(ruta_archivo)
    ws = wb[NOMBRE_HOJA]

    # Encabezados
    for celda in ws[1]:
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=False,
        )
        celda.border = BORDE_CELDA

    ws.row_dimensions[1].height = 22

    # Datos generales
    for fila in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for celda in fila:
            celda.alignment = Alignment(
                vertical="center",
                wrap_text=False,
            )
            celda.border = BORDE_CELDA

    # Anchos
    for letra, ancho in ANCHOS_COLUMNAS.items():
        ws.column_dimensions[letra].width = ancho

    # Mapa de encabezados
    encabezados = {
        ws.cell(row=1, column=columna).value: columna
        for columna in range(1, ws.max_column + 1)
    }

    # Fechas con hora
    for nombre in [
        "PRO_FECHA_SISTEMA_CREACION",
        "FECHA_LIMITE_ANS",
    ]:
        col = encabezados.get(nombre)

        if col:
            for fila in range(2, ws.max_row + 1):
                ws.cell(
                    fila,
                    col,
                ).number_format = "dd/mm/yyyy hh:mm:ss"

    # Fecha corta
    col_vencimiento = encabezados.get(
        "PRO_FECHA_VENCIMIENTO"
    )

    if col_vencimiento:
        for fila in range(2, ws.max_row + 1):
            ws.cell(
                fila,
                col_vencimiento,
            ).number_format = "dd/mm/yyyy"

    # Enteros
    for nombre in [
        "DIAS_CONTRACTUALES",
        "DIAS_TRANSCURRIDOS",
        "DIAS_PARA_INICIAR_ALERTA",
        "DIAS_RESTANTES",
    ]:
        col = encabezados.get(nombre)

        if col:
            for fila in range(2, ws.max_row + 1):
                ws.cell(
                    fila,
                    col,
                ).number_format = "0"

    # ========================================================
    # COLORES DIRECTOS EN LA COLUMNA ESTADO
    # ========================================================

    col_estado = encabezados.get("ESTADO")

    if col_estado:
        for fila in range(2, ws.max_row + 1):
            celda_estado = ws.cell(
                row=fila,
                column=col_estado,
            )

            estado = str(
                celda_estado.value or ""
            ).strip().upper()

            if estado == "ALERTA":
                celda_estado.fill = RELLENO_ALERTA

            elif estado == "VENCIDO":
                celda_estado.fill = RELLENO_VENCIDO

            elif estado == "ALERTA 0 DIAS":
                celda_estado.fill = RELLENO_ALERTA_CERO

            elif estado == "A TIEMPO":
                celda_estado.fill = RELLENO_A_TIEMPO

            celda_estado.font = FUENTE_ESTADO
            celda_estado.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )

    # Vista
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(ruta_archivo)
    wb.close()


# ============================================================
# GENERACIÓN
# ============================================================

def generar_informe_ans_redes() -> Path:
    CARPETA_SALIDA_REDES.mkdir(
        parents=True,
        exist_ok=True,
    )

    df, archivo_origen = procesar_exporte_redes()

    faltantes = [
        columna
        for columna in COLUMNAS_SALIDA
        if columna not in df.columns
    ]

    if faltantes:
        raise ValueError(
            "No se puede generar INFORME_ANS_REDES.xlsx porque faltan columnas:\n"
            + "\n".join(
                f" - {columna}"
                for columna in faltantes
            )
        )

    df_salida = _preparar_fechas(
        df[COLUMNAS_SALIDA].copy()
    )

    ruta_salida = (
        CARPETA_SALIDA_REDES
        / NOMBRE_ARCHIVO_SALIDA
    )

    try:
        with pd.ExcelWriter(
            ruta_salida,
            engine="openpyxl",
            mode="w",
            datetime_format="dd/mm/yyyy hh:mm:ss",
            date_format="dd/mm/yyyy",
        ) as writer:
            df_salida.to_excel(
                writer,
                sheet_name=NOMBRE_HOJA,
                index=False,
            )

    except PermissionError as error:
        raise PermissionError(
            f"No fue posible actualizar:\n{ruta_salida}\n\n"
            "Cierre INFORME_ANS_REDES.xlsx en Excel y vuelva a ejecutar."
        ) from error

    _aplicar_formato(
        ruta_salida
    )

    print("=" * 70)
    print("INFORME ANS REDES GENERADO CORRECTAMENTE")
    print("=" * 70)
    print(f"Archivo origen      : {archivo_origen.name}")
    print(f"Registros generados : {len(df_salida):,}")
    print(f"Columnas generadas  : {len(df_salida.columns)}")
    print(f"Hoja                : {NOMBRE_HOJA}")
    print(f"Archivo de salida   : {ruta_salida}")
    print()
    print("COLORES ESTADO:")
    print(" - ALERTA        -> AMARILLO")
    print(" - VENCIDO       -> ROJO")
    print(" - ALERTA 0 DIAS -> NARANJA")
    print(" - A TIEMPO      -> VERDE")

    return ruta_salida


if __name__ == "__main__":
    try:
        generar_informe_ans_redes()

    except Exception as error:
        print("=" * 70)
        print("ERROR AL GENERAR INFORME ANS REDES")
        print("=" * 70)
        print(error)
        raise SystemExit(1)
